import os
import openpyxl
import warnings
import requests

def average(n1: int | float, n2: int | float, n3: int | float) -> float:
    """
   gets the average between 3 numbers

   Parameters
   ----------
      n1 & n2 & n3: float | int
          defines the numbers to average

   Returns
   -------
      float
          the result of averaging the above parameters
   """
    return (n1 + n2 + n3)/3

print(average(7, 5, 9))
print(average(6, 6, 7))
print()

# the commented code below will not run and will error if the above code
# were replaced. This is because python is an interpreted language and reads
# files top to bottom, so when the interpreter tries to run the function in the
# print statement, it wouldn't have seen the function's declaration yet and can't
# execute the function. (This is also why `if __name__ == "__main__"` has to be
# at the bottom of the file)


# print(average(7, 5, 9))
# print(average(6, 6, 7))

# def average(n1: int | float, n2: int | float, n3: int | float) -> float:
#    """
#    gets the average between 3 numbers
#
#    Parameters
#    ----------
#       n1 & n2 & n3: float | int
#           defines the numbers to average
#
#    Returns
#    -------
#       float
#           the result of averaging the above parameters
#    """
#     return (n1 + n2 + n3)/3


# Errors because function parameters are local to the function itself and can
# only be accessed from inside the function. Parameters can be thought as input
# placeholders to where they will be used in the function:
#
# print(n1)

def dog_to_human_years(dog_years: int) -> int:
    """
    the conversion between dog years and human years can be represented by this
    formula: dog's age in human years = 24 + (dog's age - 2) x 4

    Parameters
    ----------
        dog_years: int
            the age of the dog

    Returns
    -------
        int
            the age of the dog in human years
    """
    return (24 + (dog_years - 2)) * 4

def get_car_value(purchase_price: float, age_of_car: int, type_of_car: str):
    """
    calculates the current value of a car based on the type of the car and the
    age of the car.

    Parameters
    ----------
        purchase_price : float
            defines the price of the car when first bought
        age_of_car : int
            defines the number of years after the car was first bought
        type_of_car : str
            defines the type of car to get its appreciation/depreciation rate

    Returns
        None
            prints out the value of the type of car after n years
    -------
    """

    # `str.casefold()` converts all strings to lower case
    type_of_car_lower = type_of_car.casefold()

    current_value: float;

    match type_of_car_lower:
        case "german":
            current_value = purchase_price * (1 - (0.05 * age_of_car))
        case "japanese":
            current_value = purchase_price * (1 - (0.07 * age_of_car))
        case "italian":
            current_value = purchase_price * (1 + (0.05 * age_of_car))
        case _:
            print("Unknown car type:", type_of_car_lower)
            return

    if type_of_car_lower[0] == "i":
        print(f"The value of an {type_of_car_lower} after {age_of_car} years will be ${current_value}")
    else:
        print(f"The value of a {type_of_car_lower} after {age_of_car} years will be ${current_value}")

def dog_questions():
    """
    asks the user the name and age of their dog and prints out their dog's
    age in human years

    Parameters
    ----------
        None

    Returns
    -------
        None
    """
    name = input("What is your dog's name? ")

    dog_years = input(f"How old is {name}? ")

    dog_years = int(dog_years)

    human_years = dog_to_human_years(dog_years)
    print(f"Your {name} is {human_years} years old.")


def ice_cream_price():
    """
    asks the user the number of scoops of ice cream they want, calculates
    the price by this formula: `Price = number of scoops x $1.15 + $2.25`,
    then prints out the price of an ice cream scoops with the inputted
    number of scoops

    Parameters
    ----------
        None

    Returns
    -------
        None
    """
    num_scoops = input("How many scoops of ice cream would you like? ")
    num_scoops = int(num_scoops)

    cone_price = (num_scoops * 1.15) + 2.25

    print(f"A {num_scoops}-scoop cone will cost ${cone_price}")

def ensure_file_downloaded(file_path: str, url: str = "", method: str = "get", data: str = "", headers = None):
    """
    checks if a file of the given `file_path` exists, and if it doesn't it
    downloads it using the given `url`, `method`, and other request related
    prameters

    Parameters
    ----------
        file_path: str
            the file path to check and save to
        url: str
            the download url of the file
        method: str
            the request method used for downloading the file
        data: str
            the request body data to send in the request
        headers: str
            the headers to send in the request

    Returns
    -------
        None
    """

    if not os.path.exists(file_path):
        # so far we're only supporting GET and POST request mehtods
        method = method.casefold()
        resp = ""
        match method:
            case "get":
                print(f"Downloading {file_path[2:]}")
                resp = requests.request(method="get", url=url, headers=headers, data=data)
            case "post":
                print(f"Downloading {file_path[2:]}")
                resp = requests.request(method="post", url=url, headers=headers, data=data)
            case _:
                print("ERR: Unsupported Request Method")
                exit(1)
        if resp.status_code != 200:
            print("ERR: Could not download file due to Status Code", resp.status_code)
            exit(1)
        m_file = open(file_path, "wb")
        m_file.write(resp.content)
        m_file.close()

OUNCES_IN_HALF_GALLON = 64.00
OUNCES_IN_SCOOP = 4.00

def ice_cream_price_w_inflation():
    """
    same as `ice_cream_price()` except it calculates the price of each scoop
    based off of historical ice cream prices based on inflation and defining
    how much ice cream is in a scoop.
    Parameters
    ----------
        None

    Returns
    -------
        None
    """
    month_n_year = input("Input a date with the format mm/yyyy: ")
    scoops = input("Input number of ice cream scoops: ")
    scoops = int(scoops)
    month_n_year_parts = month_n_year.split("/")


    month = int(month_n_year_parts[0])
    year = int(month_n_year_parts[1])
    print(f"Month: {month}, Year: {year}")

    if not 1 <= month <= 12:
        print("Invalid month:", month)
    if year < 1980 or year > 2024:
        print("No inflation data for year:", year)

    file = open("./IceCreamFlation.xlsx", "rb")
    workbook = openpyxl.load_workbook(file, data_only=True)
    table = workbook.active
    if table == None:
        print("ERR: No active table in the workbook")
        return

    max_row = int(table.max_row) if table.max_row.is_integer() else 0
    max_column = int(table.max_column) if table.max_column.is_integer() else 0
    if max_row == 0 or max_column == 0:
        print("ERR: No data in table")
        return

    selected_cell = table.cell(
        row=max_row - (2024 - year),
        column=month+1
    )
    price_per_half_gallon = 0.00
    try:
        price_per_half_gallon = float(selected_cell.value)
    except:
        print("Yeah idk")
        exit(1)

    price_per_ounce = float(price_per_half_gallon)/OUNCES_IN_HALF_GALLON
    price_per_scoop = price_per_ounce*4
    print(f"{scoops} scoops of ice cream in {month}/{year}, would cost about ${price_per_scoop*scoops}")

if __name__ == "__main__":
    url = "https://data.bls.gov/pdq/SurveyOutputServlet"

    # It was actually quite tricky to get this request data since I had to
    # capture the request made when downloading the file from the site normally.
    # And usually you're not able to do this because the download 'link' on
    # the page is represented as a form element with several hidden values to
    # be used in the request. Specifically, the problem with this form though is
    # that it opens up a new tab to download the file and immediately closes it
    # right after, not letting us easily capture the request data (granted I
    # could've also used something like WireShark which is a lot better suited
    # for this kind of thing, but I think the way I did this is more fun).
    #
    # <form name="excel" action="/pdq/SurveyOutputServlet" method="POST" target="_blank">
    #                                                                    ^
    #            This is caused by this `target="_blank"` attribute here-|
    # However, since we're in a web browser, we're able to easily view and edit
    # the page contents to our will!
    #
    # So after going into the DevTools by right clicking and inspecting the
    # element, we're able to just remove the problematic form attribute:
    #
    # <form name="excel" action="/pdq/SurveyOutputServlet" method="POST">
    #
    # Now when we go ahead and click on the element, instead of opening up a
    # new tab and closing it immediately, the download happens within the
    # current open tab, which stays open after the request, allowing us to
    # catch the request in the Network tab in the DevTools, revealing the
    # request's data in full!
    # Which the necessary can be viewed here (broke up the url encoded values
    # into pieces for readability):
    data="".join([
        "x=31&",
        "y=13&",
        "request_action=get_data&",
        "reformat=true&",
        "from_results_page=true&",
        "years_option=specific_years&",
        "delimiter=comma&",
        "output_type=multi&",
        "periods_option=all_periods&",
        "output_view=data&",
        "to_year=2024&",
        "from_year=1980&",
        "output_format=excelTable&",
        "original_output_type=default&",
        "annualAveragesRequested=false&",
        "series_id=APU0000710411"
    ])
    ensure_file_downloaded("./IceCreamFlation.xlsx", url, "post", data, {
        "content-type": "application/x-www-form-urlencoded"
    })

    dog_years = 5
    print(f"{dog_years} dog years is equivalent to {dog_to_human_years(dog_years)} human years")

    dog_years = 11
    print(f"{dog_years} dog years is equivalent to {dog_to_human_years(dog_years)} human years")

    get_car_value(30000, 5, "GERMAN")
    get_car_value(30000, 5, "Japanese")
    get_car_value(30000, 5, "iTALIAN")
    get_car_value(30000, 5, "polish")

    dog_questions()
    ice_cream_price()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        ice_cream_price_w_inflation()
